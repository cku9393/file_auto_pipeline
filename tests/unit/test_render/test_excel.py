"""
test_excel.py - Excel (XLSX) 렌더러 테스트

ADR-0002 검증:
- Named Range > cell_address 우선순위
- 둘 다 있으면 에러 (fail-fast)
- 측정 테이블: start_row 기반
"""

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from src.domain.errors import ErrorCodes, PolicyRejectError
from src.render.excel import ExcelRenderer, load_manifest, render_xlsx


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def simple_xlsx_template(tmp_path: Path) -> Path:
    """
    간단한 XLSX 템플릿 생성.

    - Named Range: WO_NO (B2), LINE (B3)
    - 일반 셀: B4 (result)
    """
    template_path = tmp_path / "template.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 라벨
    ws["A2"] = "작업번호"
    ws["A3"] = "라인"
    ws["A4"] = "결과"

    # Named Range 정의
    # openpyxl에서 Named Range 추가
    wb.defined_names.add(DefinedName("WO_NO", attr_text="Sheet1!$B$2"))
    wb.defined_names.add(DefinedName("LINE", attr_text="Sheet1!$B$3"))

    wb.save(template_path)

    return template_path


@pytest.fixture
def xlsx_with_measurements_template(tmp_path: Path) -> Path:
    """
    측정 데이터 테이블이 있는 XLSX 템플릿.

    - 헤더: Row 4 (A: 항목, B: 규격, C: 측정값, D: 결과)
    - 데이터: Row 5부터
    """
    template_path = tmp_path / "template_measurements.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "검사"

    # 헤더 정보
    ws["A1"] = "작업번호"
    ws["B1"] = ""  # 값은 렌더링 시 채움

    # 측정 테이블 헤더
    ws["A4"] = "항목"
    ws["B4"] = "규격"
    ws["C4"] = "측정값"
    ws["D4"] = "결과"

    # Named Range
    wb.defined_names.add(DefinedName("WO_NO", attr_text="검사!$B$1"))

    wb.save(template_path)

    return template_path


@pytest.fixture
def sample_manifest() -> dict:
    """기본 manifest."""
    return {
        "template_id": "test_template",
        "xlsx_mappings": {
            "named_ranges": {
                "wo_no": "WO_NO",
                "line": "LINE",
            },
            "cell_addresses": {
                "result": "Sheet1!B4",
            },
            "measurements": {
                "sheet": "Sheet1",
                "start_row": 5,
                "columns": {
                    "item": "A",
                    "spec": "B",
                    "measured": "C",
                    "result": "D",
                },
            },
        },
    }


@pytest.fixture
def conflicting_manifest() -> dict:
    """충돌하는 manifest (named_range + cell_address 동일 필드)."""
    return {
        "template_id": "conflict",
        "xlsx_mappings": {
            "named_ranges": {
                "wo_no": "WO_NO",  # wo_no 있음
            },
            "cell_addresses": {
                "wo_no": "Sheet1!B2",  # wo_no 또 있음 → 충돌!
            },
        },
    }


@pytest.fixture
def sample_data() -> dict:
    """렌더링 테스트용 데이터."""
    return {
        "wo_no": "WO-001",
        "line": "L1",
        "result": "PASS",
        "measurements": [
            {"item": "길이", "spec": "10±0.1", "measured": Decimal("10.05"), "result": "PASS"},
            {"item": "폭", "spec": "5±0.1", "measured": Decimal("5.02"), "result": "PASS"},
        ],
    }


# =============================================================================
# ExcelRenderer 초기화 테스트
# =============================================================================

class TestExcelRendererInit:
    """ExcelRenderer 초기화 테스트."""

    def test_init_with_valid_template(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
    ):
        """유효한 템플릿과 manifest로 초기화."""
        renderer = ExcelRenderer(simple_xlsx_template, sample_manifest)

        assert renderer.template_path == simple_xlsx_template
        assert renderer.manifest == sample_manifest

    def test_init_with_nonexistent_template(
        self,
        tmp_path: Path,
        sample_manifest: dict,
    ):
        """존재하지 않는 템플릿 → PolicyRejectError."""
        nonexistent = tmp_path / "nonexistent.xlsx"

        with pytest.raises(PolicyRejectError) as exc_info:
            ExcelRenderer(nonexistent, sample_manifest)

        assert exc_info.value.code == ErrorCodes.TEMPLATE_NOT_FOUND

    def test_init_with_conflicting_manifest(
        self,
        simple_xlsx_template: Path,
        conflicting_manifest: dict,
    ):
        """
        ADR-0002: named_range + cell_address 동일 필드 → 에러 (fail-fast).
        """
        with pytest.raises(PolicyRejectError) as exc_info:
            ExcelRenderer(simple_xlsx_template, conflicting_manifest)

        assert exc_info.value.code == ErrorCodes.RENDER_FAILED
        assert "conflict" in exc_info.value.context.get("error", "").lower()
        assert "wo_no" in exc_info.value.context.get("fields", [])


# =============================================================================
# ExcelRenderer.render 테스트
# =============================================================================

class TestExcelRendererRender:
    """ExcelRenderer.render 테스트."""

    def test_basic_render(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
        sample_data: dict,
        tmp_path: Path,
    ):
        """기본 렌더링."""
        renderer = ExcelRenderer(simple_xlsx_template, sample_manifest)
        output_path = tmp_path / "output.xlsx"

        result = renderer.render(sample_data, output_path)

        assert result == output_path
        assert output_path.exists()

        # 렌더링 결과 확인
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        # Named Range로 채워진 값 확인
        assert ws["B2"].value == "WO-001"  # WO_NO named range
        assert ws["B3"].value == "L1"       # LINE named range

        # Cell Address로 채워진 값 확인
        assert ws["B4"].value == "PASS"

    def test_creates_output_directory(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
        sample_data: dict,
        tmp_path: Path,
    ):
        """출력 디렉터리 자동 생성."""
        renderer = ExcelRenderer(simple_xlsx_template, sample_manifest)
        output_path = tmp_path / "nested" / "dir" / "output.xlsx"

        renderer.render(sample_data, output_path)

        assert output_path.exists()
        assert output_path.parent.exists()

    def test_named_range_priority(
        self,
        simple_xlsx_template: Path,
        sample_data: dict,
        tmp_path: Path,
    ):
        """
        ADR-0002: Named Range 우선.

        named_ranges와 cell_addresses에 서로 다른 필드가 있을 때
        각각 올바르게 적용되는지 확인.
        """
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {
                    "wo_no": "WO_NO",  # Named Range 사용
                },
                "cell_addresses": {
                    "line": "Sheet1!B3",  # Cell Address 사용
                },
            },
        }

        renderer = ExcelRenderer(simple_xlsx_template, manifest)
        output_path = tmp_path / "output.xlsx"

        renderer.render(sample_data, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        assert ws["B2"].value == "WO-001"  # Named Range
        assert ws["B3"].value == "L1"       # Cell Address

    def test_measurements_fill(
        self,
        xlsx_with_measurements_template: Path,
        tmp_path: Path,
    ):
        """
        ADR-0002: start_row 기반 측정 테이블 채우기.
        """
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {"wo_no": "WO_NO"},
                "cell_addresses": {},
                "measurements": {
                    "sheet": "검사",
                    "start_row": 5,
                    "columns": {
                        "item": "A",
                        "spec": "B",
                        "measured": "C",
                        "result": "D",
                    },
                },
            },
        }

        data = {
            "wo_no": "WO-001",
            "measurements": [
                {"item": "길이", "spec": "10±0.1", "measured": Decimal("10.05"), "result": "PASS"},
                {"item": "폭", "spec": "5±0.1", "measured": Decimal("5.02"), "result": "PASS"},
            ],
        }

        renderer = ExcelRenderer(xlsx_with_measurements_template, manifest)
        output_path = tmp_path / "output.xlsx"

        renderer.render(data, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb["검사"]

        # Row 5 (첫 번째 측정 행)
        assert ws["A5"].value == "길이"
        assert ws["B5"].value == "10±0.1"
        assert ws["C5"].value == 10.05  # Decimal → float
        assert ws["D5"].value == "PASS"

        # Row 6 (두 번째 측정 행)
        assert ws["A6"].value == "폭"
        assert ws["B6"].value == "5±0.1"
        assert ws["C6"].value == 5.02
        assert ws["D6"].value == "PASS"

    def test_decimal_to_float_conversion(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
        tmp_path: Path,
    ):
        """Decimal 값은 float로 변환."""
        data = {
            "wo_no": Decimal("123.456"),
            "line": "L1",
        }

        renderer = ExcelRenderer(simple_xlsx_template, sample_manifest)
        output_path = tmp_path / "output.xlsx"

        renderer.render(data, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        assert ws["B2"].value == 123.456
        assert isinstance(ws["B2"].value, float)

    def test_missing_named_range_ignored(
        self,
        simple_xlsx_template: Path,
        tmp_path: Path,
    ):
        """정의되지 않은 Named Range는 무시 (에러 아님)."""
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {
                    "wo_no": "NONEXISTENT_RANGE",  # 템플릿에 없음
                },
                "cell_addresses": {},
            },
        }

        data = {"wo_no": "WO-001"}

        renderer = ExcelRenderer(simple_xlsx_template, manifest)
        output_path = tmp_path / "output.xlsx"

        # 에러 없이 렌더링 완료 (값은 채워지지 않음)
        renderer.render(data, output_path)
        assert output_path.exists()

    def test_render_with_korean(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
        tmp_path: Path,
    ):
        """한글 데이터 렌더링."""
        data = {
            "wo_no": "작업-001",
            "line": "라인A",
            "result": "합격",
        }

        renderer = ExcelRenderer(simple_xlsx_template, sample_manifest)
        output_path = tmp_path / "output.xlsx"

        renderer.render(data, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active

        assert ws["B2"].value == "작업-001"
        assert ws["B3"].value == "라인A"
        assert ws["B4"].value == "합격"


# =============================================================================
# _validate_manifest 테스트 (충돌 검증)
# =============================================================================

class TestManifestValidation:
    """manifest 검증 테스트."""

    def test_no_conflict_passes(
        self,
        simple_xlsx_template: Path,
    ):
        """충돌 없는 manifest는 통과."""
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {"wo_no": "WO_NO"},
                "cell_addresses": {"result": "B4"},
            },
        }

        # 에러 없이 생성
        renderer = ExcelRenderer(simple_xlsx_template, manifest)
        assert renderer is not None

    def test_conflict_raises_error(
        self,
        simple_xlsx_template: Path,
    ):
        """동일 필드가 양쪽에 있으면 에러."""
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {"wo_no": "WO_NO", "line": "LINE"},
                "cell_addresses": {"line": "B3"},  # line 중복!
            },
        }

        with pytest.raises(PolicyRejectError) as exc_info:
            ExcelRenderer(simple_xlsx_template, manifest)

        assert exc_info.value.code == ErrorCodes.RENDER_FAILED
        assert "line" in exc_info.value.context.get("fields", [])

    def test_multiple_conflicts(
        self,
        simple_xlsx_template: Path,
    ):
        """여러 필드 충돌 시 모두 보고."""
        manifest = {
            "xlsx_mappings": {
                "named_ranges": {"wo_no": "WO_NO", "line": "LINE"},
                "cell_addresses": {"wo_no": "B2", "line": "B3"},  # 둘 다 중복
            },
        }

        with pytest.raises(PolicyRejectError) as exc_info:
            ExcelRenderer(simple_xlsx_template, manifest)

        fields = exc_info.value.context.get("fields", [])
        assert "wo_no" in fields
        assert "line" in fields


# =============================================================================
# render_xlsx 간편 함수 테스트
# =============================================================================

class TestRenderXlsx:
    """render_xlsx 간편 함수 테스트."""

    def test_basic_usage(
        self,
        simple_xlsx_template: Path,
        sample_manifest: dict,
        sample_data: dict,
        tmp_path: Path,
    ):
        """기본 사용."""
        output_path = tmp_path / "output.xlsx"

        result = render_xlsx(
            simple_xlsx_template,
            sample_manifest,
            sample_data,
            output_path,
        )

        assert result == output_path
        assert output_path.exists()


# =============================================================================
# load_manifest 테스트
# =============================================================================

class TestLoadManifest:
    """load_manifest 함수 테스트."""

    def test_loads_yaml(self, tmp_path: Path):
        """YAML 파일 로드."""
        manifest_path = tmp_path / "manifest.yaml"
        manifest_path.write_text(
            """
template_id: test
xlsx_mappings:
  named_ranges:
    wo_no: WO_NO
""",
            encoding="utf-8",
        )

        result = load_manifest(manifest_path)

        assert result["template_id"] == "test"
        assert result["xlsx_mappings"]["named_ranges"]["wo_no"] == "WO_NO"

    def test_handles_korean(self, tmp_path: Path):
        """한글 포함 YAML."""
        manifest_path = tmp_path / "manifest.yaml"
        manifest_path.write_text(
            """
display_name: 고객사A 검사성적서
""",
            encoding="utf-8",
        )

        result = load_manifest(manifest_path)

        assert result["display_name"] == "고객사A 검사성적서"
