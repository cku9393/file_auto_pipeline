"""
Excel (XLSX) 렌더러: openpyxl 기반.

spec-v2.md Section 9.2 + ADR-0002 참조:
- 템플릿 복사 후 값 채우기
- Named Range 우선 (권장)
- cell_addresses는 named_range 없을 때만
- 둘 다 있으면 에러 (fail-fast)
- 측정 테이블: start_row 기반
"""

from decimal import Decimal
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.domain.errors import ErrorCodes, PolicyRejectError


class ExcelRenderer:
    """
    Excel 문서 렌더러.

    Usage:
        renderer = ExcelRenderer(template_path, manifest)
        renderer.render(data, output_path)
    """

    def __init__(self, template_path: Path, manifest: dict[str, Any]):
        """
        Args:
            template_path: XLSX 템플릿 파일 경로
            manifest: manifest.yaml 내용 (xlsx_mappings 포함)

        Raises:
            PolicyRejectError: TEMPLATE_NOT_FOUND
        """
        if not template_path.exists():
            raise PolicyRejectError(
                ErrorCodes.TEMPLATE_NOT_FOUND,
                path=str(template_path),
            )

        self.template_path = template_path
        self.manifest = manifest
        self._validate_manifest()

    def _validate_manifest(self) -> None:
        """
        manifest 검증: 동일 필드에 named_range + cell_address 둘 다 있으면 에러.

        ADR-0002: 둘 다 있으면 fail-fast
        """
        mappings = self.manifest.get("xlsx_mappings", {})
        named_ranges = set(mappings.get("named_ranges", {}).keys())
        cell_addresses = set(mappings.get("cell_addresses", {}).keys())

        # 교집합 확인
        conflicts = named_ranges & cell_addresses
        if conflicts:
            raise PolicyRejectError(
                ErrorCodes.RENDER_FAILED,
                error="XLSX mapping conflict: same field in both named_ranges and cell_addresses",
                fields=list(conflicts),
            )

    def render(
        self,
        data: dict[str, Any],
        output_path: Path,
    ) -> Path:
        """
        템플릿에 데이터를 채워 Excel 문서 생성.

        Args:
            data: 템플릿에 채울 데이터
                - wo_no, line, part_no, lot, result 등
                - measurements: list[dict] (측정 데이터)
            output_path: 출력 파일 경로

        Returns:
            저장된 파일 경로

        Raises:
            PolicyRejectError: RENDER_FAILED
        """
        try:
            # 템플릿 복사
            wb = load_workbook(self.template_path)

            # 필드 채우기
            self._fill_fields(wb, data)

            # 측정 데이터 채우기
            self._fill_measurements(wb, data.get("measurements", []))

            # 저장
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

            return output_path

        except PolicyRejectError:
            raise
        except Exception as e:
            raise PolicyRejectError(
                ErrorCodes.RENDER_FAILED,
                template=str(self.template_path),
                error=str(e),
            ) from e

    def _fill_fields(self, wb: Workbook, data: dict[str, Any]) -> None:
        """필드 값 채우기 (Named Range 우선)."""
        mappings = self.manifest.get("xlsx_mappings", {})
        named_ranges = mappings.get("named_ranges", {})
        cell_addresses = mappings.get("cell_addresses", {})

        # 모든 필드 키 수집
        all_fields = set(named_ranges.keys()) | set(cell_addresses.keys())

        for field in all_fields:
            value = data.get(field)
            if value is None:
                continue

            # 우선순위 1: Named Range
            if field in named_ranges:
                self._set_named_range_value(wb, named_ranges[field], value)
            # 우선순위 2: Cell Address
            elif field in cell_addresses:
                self._set_cell_value(wb, cell_addresses[field], value)

    def _set_named_range_value(
        self,
        wb: Workbook,
        range_name: str,
        value: Any,
    ) -> None:
        """Named Range에 값 설정."""
        if range_name not in wb.defined_names:
            # Named Range가 없으면 경고만 (fail-fast 대신 유연하게)
            return

        # Named Range의 목적지 가져오기
        destinations = wb.defined_names[range_name].destinations
        for sheet_name, cell_ref in destinations:
            ws = wb[sheet_name]
            # cell_ref가 범위일 수 있음 (예: $A$1:$A$1)
            # 첫 번째 셀만 사용
            cell_addr = cell_ref.replace("$", "").split(":")[0]
            ws[cell_addr] = self._convert_value(value)

    def _set_cell_value(
        self,
        wb: Workbook,
        cell_address: str,
        value: Any,
    ) -> None:
        """직접 셀 주소에 값 설정."""
        # 형식: "Sheet1!B4" 또는 "B4" (기본 시트)
        if "!" in cell_address:
            sheet_name, cell_ref = cell_address.split("!", 1)
            ws = wb[sheet_name]
        else:
            ws = wb.active
            cell_ref = cell_address

        ws[cell_ref] = self._convert_value(value)

    def _fill_measurements(
        self,
        wb: Workbook,
        measurements: list[dict[str, Any]],
    ) -> None:
        """측정 데이터 채우기 (start_row 기반)."""
        mappings = self.manifest.get("xlsx_mappings", {})
        meas_config = mappings.get("measurements", {})

        if not meas_config:
            return

        sheet_name = meas_config.get("sheet", wb.active.title)
        start_row = meas_config.get("start_row", 5)
        columns = meas_config.get("columns", {})

        try:
            ws = wb[sheet_name]
        except KeyError:
            ws = wb.active

        for i, row_data in enumerate(measurements):
            row_num = start_row + i

            for field, col_letter in columns.items():
                value = row_data.get(field)
                if value is not None:
                    ws[f"{col_letter}{row_num}"] = self._convert_value(value)

    def _convert_value(self, value: Any) -> Any:
        """값 변환 (Decimal → float 등)."""
        if isinstance(value, Decimal):
            # Excel은 Decimal을 직접 지원하지 않음
            # float로 변환하되, 문자열이 더 안전할 수 있음
            return float(value)
        return value


def render_xlsx(
    template_path: Path,
    manifest: dict[str, Any],
    data: dict[str, Any],
    output_path: Path,
) -> Path:
    """
    Excel 문서 생성 (간편 함수).

    Args:
        template_path: XLSX 템플릿 파일 경로
        manifest: manifest.yaml 내용
        data: 템플릿에 채울 데이터
        output_path: 출력 파일 경로

    Returns:
        저장된 파일 경로
    """
    renderer = ExcelRenderer(template_path, manifest)
    return renderer.render(data, output_path)


def load_manifest(manifest_path: Path) -> dict[str, Any]:
    """
    manifest.yaml 로드.

    Args:
        manifest_path: manifest.yaml 경로

    Returns:
        manifest 내용
    """
    with open(manifest_path, encoding="utf-8") as f:
        data: dict[str, Any] = yaml.safe_load(f)
        return data
