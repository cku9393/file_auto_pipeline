"""
XLSX semantic extractor for golden tests.

Extracts meaningful content from XLSX files into a normalized JSON structure,
avoiding brittle binary comparisons.

Extracted elements:
1. Sheet names
2. Cell values (specified ranges or key cells)
3. Measurement table data

Normalization:
- Numbers converted to normalized string representation
- Empty cells → null
- Formulas → evaluated result (not formula text)
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .normalize import Normalizer


@dataclass
class XlsxContent:
    """Extracted XLSX content structure."""

    sheets: list[str] = field(default_factory=list)
    cells: dict[str, dict[str, Any]] = field(default_factory=dict)
    measurements: list[dict[str, Any]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            "sheets": self.sheets,
            "cells": self.cells,
            "measurements": self.measurements,
            "metadata": self.metadata,
        }


class XlsxExtractor:
    """
    Extract semantic content from XLSX files.

    Usage:
        extractor = XlsxExtractor()
        content = extractor.extract(Path("spreadsheet.xlsx"))
        json_data = content.to_dict()

    Or with specific cell extraction:
        extractor = XlsxExtractor(
            key_cells={"Sheet1": ["A1", "B2", "C3"]},
            cell_ranges={"Sheet1": "A1:E10"},
        )
    """

    def __init__(
        self,
        normalizer: Normalizer | None = None,
        key_cells: dict[str, list[str]] | None = None,
        cell_ranges: dict[str, str] | None = None,
        measurement_config: dict[str, Any] | None = None,
        extract_all_cells: bool = False,
    ):
        """
        Args:
            normalizer: Normalizer instance for value processing
            key_cells: Specific cells to extract per sheet {"Sheet1": ["A1", "B2"]}
            cell_ranges: Cell ranges to extract per sheet {"Sheet1": "A1:E10"}
            measurement_config: Measurement table config (from manifest)
            extract_all_cells: Extract all non-empty cells (can be large)
        """
        self.normalizer = normalizer or Normalizer()
        self.key_cells = key_cells or {}
        self.cell_ranges = cell_ranges or {}
        self.measurement_config = measurement_config or {}
        self.extract_all_cells = extract_all_cells

    def extract(self, xlsx_path: Path) -> XlsxContent:
        """
        Extract content from an XLSX file.

        Args:
            xlsx_path: Path to the XLSX file

        Returns:
            XlsxContent with extracted elements
        """
        wb = load_workbook(xlsx_path, data_only=True)
        content = XlsxContent()

        # Extract sheet names
        content.sheets = wb.sheetnames

        # Extract cells
        content.cells = self._extract_cells(wb)

        # Extract measurements if configured
        if self.measurement_config:
            content.measurements = self._extract_measurements(wb)

        # Extract metadata
        content.metadata = self._extract_metadata(wb)

        wb.close()
        return content

    def _extract_cells(self, wb: Workbook) -> dict[str, dict[str, Any]]:
        """
        Extract cell values.

        Returns: {"Sheet1": {"A1": value, "B2": value, ...}, ...}
        """
        cells: dict[str, dict[str, Any]] = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_cells: dict[str, Any] = {}

            # Extract key cells for this sheet
            if sheet_name in self.key_cells:
                for cell_addr in self.key_cells[sheet_name]:
                    value = self._get_cell_value(ws, cell_addr)
                    if value is not None:
                        sheet_cells[cell_addr] = value

            # Extract cell ranges
            if sheet_name in self.cell_ranges:
                range_cells = self._extract_range(ws, self.cell_ranges[sheet_name])
                sheet_cells.update(range_cells)

            # Extract all non-empty cells if requested
            if self.extract_all_cells:
                all_cells = self._extract_all_cells(ws)
                # Merge (specific cells take precedence)
                all_cells.update(sheet_cells)
                sheet_cells = all_cells

            if sheet_cells:
                cells[sheet_name] = sheet_cells

        return cells

    def _get_cell_value(self, ws: Worksheet, cell_addr: str) -> Any:
        """Get and normalize a cell value."""
        try:
            cell = ws[cell_addr]
            return self._normalize_cell_value(cell)
        except Exception:
            return None

    def _normalize_cell_value(self, cell: Cell) -> Any:
        """Normalize a cell's value for comparison."""
        value = cell.value

        if value is None:
            return None

        # Handle numbers
        if isinstance(value, (int, float)):
            return self.normalizer._normalize_number(value)

        # Handle strings
        if isinstance(value, str):
            normalized = self.normalizer._normalize_string(value)
            return normalized if normalized else None

        # Handle other types (dates, etc.)
        return str(value)

    def _extract_range(
        self,
        ws: Worksheet,
        range_str: str,
    ) -> dict[str, Any]:
        """Extract cells from a range (e.g., "A1:E10")."""
        cells: dict[str, Any] = {}

        try:
            for row in ws[range_str]:
                for cell in row:
                    value = self._normalize_cell_value(cell)
                    if value is not None:
                        cells[cell.coordinate] = value
        except Exception:
            pass

        return cells

    def _extract_all_cells(self, ws: Worksheet) -> dict[str, Any]:
        """Extract all non-empty cells from a worksheet."""
        cells: dict[str, Any] = {}

        for row in ws.iter_rows():
            for cell in row:
                value = self._normalize_cell_value(cell)
                if value is not None:
                    cells[cell.coordinate] = value

        return cells

    def _extract_measurements(self, wb: Workbook) -> list[dict[str, Any]]:
        """
        Extract measurement table data.

        Uses measurement_config from manifest:
        {
            "sheet": "Sheet1",
            "start_row": 5,
            "columns": {"item": "A", "spec": "B", "measured": "C", ...}
        }

        Or header-based extraction (more resilient to column reordering):
        {
            "sheet": "Sheet1",
            "header_row": 4,
            "headers": {"item": "항목", "spec": "규격", "measured": "측정값", ...}
        }
        """
        measurements: list[dict[str, Any]] = []

        sheet_name = self.measurement_config.get("sheet", "Sheet1")

        try:
            ws = wb[sheet_name]
        except KeyError:
            ws = wb.active
            if ws is None:
                return measurements

        # Check if header-based extraction is configured
        if "headers" in self.measurement_config:
            return self._extract_measurements_by_header(ws)

        # Fall back to column-based extraction
        start_row = self.measurement_config.get("start_row", 1)
        columns = self.measurement_config.get("columns", {})

        if not columns:
            return measurements

        # Extract rows until we hit an empty row
        row_num = start_row
        max_empty_rows = 3  # Stop after 3 consecutive empty rows

        empty_count = 0
        while empty_count < max_empty_rows:
            row_data: dict[str, Any] = {}
            has_data = False

            for field_name, col_letter in columns.items():
                cell_addr = f"{col_letter}{row_num}"
                value = self._get_cell_value(ws, cell_addr)
                row_data[field_name] = value
                if value is not None:
                    has_data = True

            if has_data:
                measurements.append(row_data)
                empty_count = 0
            else:
                empty_count += 1

            row_num += 1

            # Safety limit
            if row_num > start_row + 1000:
                break

        return measurements

    def _extract_measurements_by_header(self, ws: Worksheet) -> list[dict[str, Any]]:
        """
        Extract measurements using header names instead of fixed columns.

        This is more resilient to column reordering in the template.

        Config example:
        {
            "header_row": 4,
            "headers": {
                "item": "항목",      # field_name: header_text
                "spec": "규격",
                "measured": "측정값",
                "result": "판정"
            }
        }
        """
        measurements: list[dict[str, Any]] = []
        header_row = self.measurement_config.get("header_row", 1)
        headers_map = self.measurement_config.get("headers", {})

        if not headers_map:
            return measurements

        # Find column indices for each header
        column_mapping: dict[str, int] = {}  # field_name -> column_index

        for cell in ws[header_row]:
            cell_value = self._normalize_cell_value(cell)
            if cell_value:
                # Check if this header matches any configured header
                for field_name, header_text in headers_map.items():
                    if str(cell_value).strip() == str(header_text).strip():
                        column_mapping[field_name] = cell.column
                        break

        if not column_mapping:
            return measurements

        # Extract data rows (starting after header)
        row_num = header_row + 1
        max_empty_rows = 3
        empty_count = 0

        while empty_count < max_empty_rows:
            row_data: dict[str, Any] = {}
            has_data = False

            for field_name, col_idx in column_mapping.items():
                cell = ws.cell(row=row_num, column=col_idx)
                value = self._normalize_cell_value(cell)
                row_data[field_name] = value
                if value is not None:
                    has_data = True

            if has_data:
                measurements.append(row_data)
                empty_count = 0
            else:
                empty_count += 1

            row_num += 1

            # Safety limit
            if row_num > header_row + 1000:
                break

        return measurements

    def _extract_metadata(self, wb: Workbook) -> dict[str, Any]:
        """Extract workbook metadata."""
        metadata: dict[str, Any] = {
            "sheet_count": len(wb.sheetnames),
        }

        # Add defined names if any
        if wb.defined_names:
            metadata["defined_names"] = list(wb.defined_names.keys())

        return metadata

    def extract_to_dict(self, xlsx_path: Path) -> dict[str, Any]:
        """
        Extract and return as dictionary.

        Convenience method for JSON serialization.
        """
        content = self.extract(xlsx_path)
        return content.to_dict()


def extract_xlsx(
    xlsx_path: Path,
    normalizer: Normalizer | None = None,
    key_cells: dict[str, list[str]] | None = None,
    measurement_config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Convenience function to extract XLSX content.

    Args:
        xlsx_path: Path to XLSX file
        normalizer: Optional normalizer instance
        key_cells: Specific cells to extract per sheet
        measurement_config: Measurement table configuration

    Returns:
        Dictionary of extracted content
    """
    extractor = XlsxExtractor(
        normalizer=normalizer,
        key_cells=key_cells,
        measurement_config=measurement_config,
    )
    return extractor.extract_to_dict(xlsx_path)
